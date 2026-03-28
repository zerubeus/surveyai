"""
Chisquare — XLSForm Parser

Parses XLSForm (.xlsx) files from Kobo Toolbox / ODK Collect.
XLSForm has three sheets: survey, choices, settings.

Returns a ParsedInstrument dataclass with structured question data.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

import structlog
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = structlog.get_logger()

# Maps XLSForm type prefixes to normalized question_type
TYPE_MAP: dict[str, str] = {
    "select_one": "single_select",
    "select_multiple": "multi_select",
    "integer": "numeric",
    "decimal": "numeric",
    "range": "numeric",
    "text": "text",
    "date": "date",
    "datetime": "date",
    "time": "date",
    "geopoint": "gps",
    "geotrace": "gps",
    "geoshape": "gps",
    "calculate": "calculated",
    "note": "note",
    "begin_group": "group",
    "end_group": "group",
    "begin_repeat": "repeat",
    "end_repeat": "repeat",
    "image": "media",
    "audio": "media",
    "video": "media",
    "file": "media",
    "barcode": "text",
    "acknowledge": "text",
    "hidden": "calculated",
    "xml-external": "calculated",
    "start": "metadata",
    "end": "metadata",
    "today": "metadata",
    "deviceid": "metadata",
    "phonenumber": "metadata",
    "username": "metadata",
    "email": "metadata",
    "audit": "metadata",
}


@dataclass
class ParsedQuestion:
    name: str
    label: dict[str, str]
    type: str
    question_type: str
    choices: list[dict[str, str | dict[str, str]]]
    relevant: str | None
    constraint: str | None
    required: bool
    hint: dict[str, str]
    appearance: str | None
    group_path: list[str]


@dataclass
class ParsedInstrument:
    title: str
    form_id: str
    version: str
    default_language: str
    languages: list[str]
    questions: list[ParsedQuestion]
    skip_logic_count: int
    has_skip_logic: bool


def _sheet_to_dicts(ws: Worksheet) -> list[dict[str, str]]:
    """Convert a worksheet to a list of dicts keyed by header row."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result: list[dict[str, str]] = []
    for row in rows[1:]:
        d: dict[str, str] = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                d[headers[i]] = str(val).strip() if val is not None else ""
        if any(v for v in d.values()):
            result.append(d)
    return result


def _extract_languages(headers: list[str]) -> list[str]:
    """Extract language names from column headers like 'label::English'."""
    langs: list[str] = []
    for h in headers:
        match = re.match(r"^label::(.+)$", h)
        if match:
            lang = match.group(1).strip()
            if lang not in langs:
                langs.append(lang)
    return langs


def _get_labels(row: dict[str, str], prefix: str, languages: list[str]) -> dict[str, str]:
    """Extract multilingual labels from a row (e.g. label::English, label::French)."""
    labels: dict[str, str] = {}
    if languages:
        for lang in languages:
            val = row.get(f"{prefix}::{lang}", "")
            if val:
                labels[lang] = val
    # Fall back to bare column (no language suffix)
    if not labels:
        val = row.get(prefix, "")
        if val:
            labels["default"] = val
    return labels


def _parse_type(raw_type: str) -> tuple[str, str, str | None]:
    """
    Parse XLSForm type column.

    Examples:
        "select_one fruit" -> ("select_one fruit", "single_select", "fruit")
        "integer" -> ("integer", "numeric", None)
        "begin_group" -> ("begin_group", "group", None)

    Returns: (original_type, normalized_question_type, list_name_or_none)
    """
    parts = raw_type.strip().split(None, 1)
    if not parts:
        return (raw_type, "text", None)

    base = parts[0].lower()
    list_name = parts[1].strip() if len(parts) > 1 else None

    question_type = TYPE_MAP.get(base, "text")
    return (raw_type.strip(), question_type, list_name)


def _parse_choices(
    choices_rows: list[dict[str, str]],
    languages: list[str],
) -> dict[str, list[dict[str, str | dict[str, str]]]]:
    """
    Parse the choices sheet into a dict keyed by list_name.

    Returns: {"fruit": [{"name": "1", "label": {"English": "Apple"}}], ...}
    """
    choice_map: dict[str, list[dict[str, str | dict[str, str]]]] = {}
    for row in choices_rows:
        list_name = row.get("list_name", "").strip()
        name = row.get("name", "").strip()
        if not list_name or not name:
            continue
        label = _get_labels(row, "label", languages)
        if list_name not in choice_map:
            choice_map[list_name] = []
        choice_map[list_name].append({"name": name, "label": label})
    return choice_map


def parse_xlsform(file_bytes: bytes) -> ParsedInstrument:
    """
    Parse an XLSForm (.xlsx) file and return structured instrument data.

    Args:
        file_bytes: Raw bytes of the .xlsx file

    Returns:
        ParsedInstrument with all questions, choices, and metadata

    Raises:
        ValueError: If the file cannot be parsed as a valid XLSForm
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Cannot open file as Excel workbook: {e}") from e

    # Get sheets (case-insensitive matching)
    sheet_names_lower = {name.lower(): name for name in wb.sheetnames}

    survey_ws = wb[sheet_names_lower["survey"]] if "survey" in sheet_names_lower else None
    choices_ws = wb[sheet_names_lower["choices"]] if "choices" in sheet_names_lower else None
    settings_ws = wb[sheet_names_lower["settings"]] if "settings" in sheet_names_lower else None

    if survey_ws is None:
        raise ValueError("XLSForm is missing required 'survey' sheet")

    # Extract survey rows and detect languages
    survey_rows = _sheet_to_dicts(survey_ws)
    survey_headers = [str(h).strip() if h is not None else "" for h in next(survey_ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
    languages = _extract_languages(survey_headers)

    # Parse choices
    choices_rows = _sheet_to_dicts(choices_ws) if choices_ws is not None else []
    # Also check choices headers for languages
    if choices_ws is not None:
        choices_headers = [str(h).strip() if h is not None else "" for h in next(choices_ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
        for lang in _extract_languages(choices_headers):
            if lang not in languages:
                languages.append(lang)

    choice_map = _parse_choices(choices_rows, languages)

    # Parse settings
    title = ""
    form_id = ""
    version = ""
    default_language = ""
    if settings_ws is not None:
        settings_rows = _sheet_to_dicts(settings_ws)
        if settings_rows:
            s = settings_rows[0]
            title = s.get("form_title", "") or s.get("title", "")
            form_id = s.get("form_id", "")
            version = s.get("version", "")
            default_language = s.get("default_language", "")

    if not default_language and languages:
        default_language = languages[0]

    # Parse survey rows into questions
    questions: list[ParsedQuestion] = []
    group_stack: list[str] = []
    skip_logic_count = 0

    for row in survey_rows:
        raw_type = row.get("type", "").strip()
        name = row.get("name", "").strip()
        if not raw_type:
            continue

        original_type, question_type, list_name = _parse_type(raw_type)
        label = _get_labels(row, "label", languages)
        hint = _get_labels(row, "hint", languages)
        relevant = row.get("relevant", "").strip() or None
        constraint = row.get("constraint", "").strip() or None
        required_str = row.get("required", "").strip().lower()
        required = required_str in ("yes", "true", "1")
        appearance = row.get("appearance", "").strip() or None

        # Track group nesting
        base_type = raw_type.strip().split(None, 1)[0].lower()
        if base_type == "begin_group" or base_type == "begin_repeat":
            group_path = list(group_stack)
            if name:
                group_stack.append(name)
        elif base_type == "end_group" or base_type == "end_repeat":
            if group_stack:
                group_stack.pop()
            group_path = list(group_stack)
        else:
            group_path = list(group_stack)

        # Get choices for select types
        choices: list[dict[str, str | dict[str, str]]] = []
        if list_name and list_name in choice_map:
            choices = choice_map[list_name]

        if relevant:
            skip_logic_count += 1

        # Use name or generate one for structural elements
        if not name:
            name = f"__{base_type}_{len(questions)}"

        questions.append(ParsedQuestion(
            name=name,
            label=label,
            type=original_type,
            question_type=question_type,
            choices=choices,
            relevant=relevant,
            constraint=constraint,
            required=required,
            hint=hint,
            appearance=appearance,
            group_path=group_path,
        ))

    wb.close()

    if not title:
        title = form_id or "Untitled Form"

    logger.info(
        "xlsform_parsed",
        question_count=len(questions),
        skip_logic_count=skip_logic_count,
        languages=languages,
    )

    return ParsedInstrument(
        title=title,
        form_id=form_id,
        version=version,
        default_language=default_language,
        languages=languages,
        questions=questions,
        skip_logic_count=skip_logic_count,
        has_skip_logic=skip_logic_count > 0,
    )
